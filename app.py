import os, io, re, json
from datetime import datetime
import anthropic
import streamlit as st
import streamlit.components.v1 as components
import pdfplumber
from docx import Document
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from reportlab.lib.enums import TA_LEFT

st.set_page_config(page_title="CV Screener – Quest", page_icon="🔍", layout="wide",
                   initial_sidebar_state="collapsed")

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Work+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

#MainMenu, footer, header { visibility: hidden; }
[data-testid="stSidebar"] { display: none; }
[data-testid="stAppViewContainer"] { background: #F4F7FA; }
* { font-family: 'Work Sans', Verdana, sans-serif !important; box-sizing: border-box; }

.block-container {
  max-width: 920px !important;
  margin: 0 auto !important;
  padding: 0 0 60px 0 !important;
  width: 90% !important;
}
@media (max-width: 768px) { .block-container { width: 96% !important; } }

p, h1, h2, h3, h4, li, label, div { text-align: left; }

/* ── Text inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea {
  border-radius: 10px !important;
  border: 1px solid #E4E9EF !important;
  font-size: 14px !important;
  font-family: 'Work Sans', sans-serif !important;
  background: #F4F7FA !important;
  color: #3A4150 !important;
  line-height: 1.5 !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
  outline: 2px solid #0075BC !important;
  outline-offset: 1px !important;
  border-color: #0075BC !important;
  box-shadow: none !important;
}

/* ── Primary button ── */
[data-testid="stButton"] button[kind="primary"] {
  background: #0075BC !important;
  color: #fff !important;
  border-radius: 11px !important;
  font-weight: 700 !important;
  font-size: 14px !important;
  padding: 11px 22px !important;
  border: none !important;
  box-shadow: 0 4px 14px rgba(0,117,188,.3) !important;
  transition: background .15s !important;
}
[data-testid="stButton"] button[kind="primary"]:hover:not(:disabled) { background: #005A91 !important; }
[data-testid="stButton"] button[kind="primary"]:disabled { opacity: 0.45 !important; cursor: not-allowed !important; }

/* ── Secondary button ── */
[data-testid="stButton"] button[kind="secondary"] {
  border-radius: 9px !important;
  font-weight: 600 !important;
  font-size: 13px !important;
  border: 1px solid #E4E9EF !important;
  background: #fff !important;
  color: #3A4150 !important;
  transition: background .15s !important;
}
[data-testid="stButton"] button[kind="secondary"]:hover:not(:disabled) { background: #F4F7FA !important; }

/* ── Shortlist active → blue ── */
button[title="undo-shortlist"] {
  background: #0075BC !important; color: #fff !important; border-color: #0075BC !important;
}
button[title="undo-shortlist"]:hover { background: #005A91 !important; border-color: #005A91 !important; }

/* ── Reject active → coral ── */
button[title="undo-reject"] {
  background: #F15A29 !important; color: #fff !important; border-color: #F15A29 !important;
}
button[title="undo-reject"]:hover { background: #c94820 !important; border-color: #c94820 !important; }

/* ── Cards ── */
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"] {
  border-radius: 14px !important;
  border: 1px solid #E4E9EF !important;
  background: #fff !important;
  box-shadow: none !important;
}
[data-testid="stVerticalBlockBorderWrapper"] > div > div { padding: 20px 22px !important; }

/* ── File uploader ── */
[data-testid="stFileUploader"] {
  border: 1.5px dashed #B8DCF0 !important;
  border-radius: 11px !important;
  background: repeating-linear-gradient(135deg,#EAF4FB 0 10px,#fff 10px 20px) !important;
}

/* ── Expander ── */
[data-testid="stExpander"] { border: 1px solid #E4E9EF !important; border-radius: 12px !important; background: #fff !important; }

/* ── Select ── */
[data-testid="stSelectbox"] > div > div { border-radius: 9px !important; border: 1px solid #E4E9EF !important; background: #fff !important; }

/* ── Candidate card hover (pure CSS) ── */
.qs-cand-card { transition: box-shadow .15s, transform .15s; }
.qs-cand-card:hover { box-shadow: 0 6px 18px rgba(30,40,60,.10); transform: translateY(-1px); }

/* ── File chip ── */
.qs-file-chip {
  display: flex; align-items: center; gap: 9px;
  background: #F4F7FA; border: 1px solid #E4E9EF;
  border-radius: 9px; padding: 9px 11px; margin-bottom: 6px;
}
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_text_from_pdf(b):
    with pdfplumber.open(io.BytesIO(b)) as p:
        return "\n".join(pg.extract_text() or "" for pg in p.pages)

def extract_text_from_docx(b):
    return "\n".join(para.text for para in Document(io.BytesIO(b)).paragraphs)

def parse_bytes(b, name):
    n = name.lower()
    if n.endswith(".pdf"):            return extract_text_from_pdf(b)
    if n.endswith((".docx", ".doc")): return extract_text_from_docx(b)
    return b.decode("utf-8", errors="replace")

def extract_file_text(f):
    try: f.seek(0)
    except: pass
    return parse_bytes(f.read(), f.name)

def initials(name):
    parts = name.split()
    return (parts[0][0] + parts[-1][0]).upper() if len(parts) > 1 else parts[0][:2].upper()

AVATAR_COLORS = ["#0075BC","#3251A3","#27AAE1","#444A67","#6B5BA8","#F7941D","#6A9A2E","#8A92A0"]

def avatar_color(name):
    return AVATAR_COLORS[sum(ord(c) for c in name) % len(AVATAR_COLORS)]

def band(score):
    if score >= 85: return "strong",   "#E3F1FA", "#005A91", "#0075BC"
    if score >= 65: return "possible", "#FEF0DC", "#C76A0A", "#F7941D"
    return "weak", "#FDE7DE", "#C23A18", "#F15A29"

def ring_color(score):
    if score >= 85: return "#0075BC"
    if score >= 65: return "#F7941D"
    return "#F15A29"

def score_ring(score, size=54):
    """Conic-gradient score ring — matches design exactly."""
    color = ring_color(score)
    inner = int(size * 0.78)
    fs    = max(11, int(size * 0.27))
    return (
        f'<div style="width:{size}px;height:{size}px;border-radius:50%;'
        f'background:conic-gradient({color} 0 {score}%,#E4E9EF {score}% 100%);'
        f'display:flex;align-items:center;justify-content:center;flex:none">'
        f'<div style="width:{inner}px;height:{inner}px;border-radius:50%;background:#fff;'
        f'display:flex;align-items:center;justify-content:center;'
        f'font-weight:800;font-size:{fs}px;color:{color};line-height:1">{score}</div>'
        f'</div>'
    )

def candidate_key(r):
    return r.get("filename") or r.get("name") or str(r.get("rank", ""))

def auto_html(content):
    st.markdown(content, unsafe_allow_html=True)

def record_history(r, action):
    key = candidate_key(r)
    if "candidate_history" not in st.session_state:
        st.session_state.candidate_history = {}
    history = st.session_state.candidate_history.setdefault(key, [])
    history.append({"action": action, "name": r.get("name", key),
                    "ts": datetime.now().strftime("%d %b %Y, %H:%M")})

def record_feedback(r, ai_overall, ai_scores, ai_labels,
                    human_overall, human_scores, reason, approved):
    key = candidate_key(r)
    fb = {"name": r.get("name", key), "role": r.get("role",""), "years": r.get("years",""),
          "ai_overall": ai_overall, "ai_scores": ai_scores,
          "human_overall": human_overall, "human_scores": human_scores,
          "competency_labels": ai_labels, "reason": reason,
          "approved": approved, "ts": datetime.now().strftime("%d %b %Y, %H:%M")}
    st.session_state.score_feedback[key] = fb
    if not approved or abs(human_overall - ai_overall) >= 5:
        example = (f"Candidate: {r.get('name','')} | Role: {r.get('role','')} | {r.get('years','')} yrs exp\n"
                   f"AI scored overall {ai_overall}/100. ")
        if approved:
            example += "Reviewer APPROVED this score."
        else:
            example += f"Reviewer ADJUSTED overall to {human_overall}/100. Reason: {reason}. "
            deltas = [f"{l}: AI={a} → Reviewer={h}"
                      for l, a, h in zip(ai_labels, ai_scores, human_scores) if abs(h-a) >= 5]
            if deltas: example += "Competency adjustments: " + "; ".join(deltas) + "."
        st.session_state.feedback_examples.append(example)

# ── Claude API ────────────────────────────────────────────────────────────────

def build_prompt(jd, competencies, cvs):
    cv_block = "".join(f"\n---\nCV #{i} – {name}\n{text}\n"
                       for i, (name, text) in enumerate(cvs.items(), 1))
    examples = st.session_state.get("feedback_examples", [])
    calibration_block = ""
    if examples:
        calibration_block = (
            "\n## Reviewer Calibration (learn from these past adjustments)\n"
            "Use these to calibrate your scoring for this batch.\n"
            + "\n".join(f"- {ex}" for ex in examples[-10:]) + "\n"
        )
    return f"""You are an expert HR screener for Quest Alliance, an NGO focused on youth skilling in India.

## Job Description
{jd}

## Required Skill Competencies
{competencies or "(derive from JD)"}
{calibration_block}
## Candidate CVs
{cv_block}

## Task
Review each candidate carefully and return a JSON array sorted best-to-worst. Each object MUST have:
- "rank": integer from 1
- "filename": CV filename exactly as given
- "name": candidate's full name (extract from CV)
- "role": their current/most recent role title
- "years": integer years of total experience
- "location": city/state from CV
- "email": email address if present, else ""
- "phone": phone if present, else ""
- "overall": integer 0-100 match score
- "scores": array of 5 integers (one per competency, same order as competencies)
- "competency_labels": array of 5 strings (the competency names scored)
- "shortlisted": true if overall >= 65
- "summary": 2-3 sentence AI summary of the candidate's fit
- "strengths": array of 2-4 short strength strings
- "gaps": array of 1-2 gap strings
- "flag": a one-sentence thing to verify, or null
- "evidence": array of 2-3 objects with "label" (competency name, uppercase) and "text" (direct quote or paraphrase from CV)

Return ONLY the JSON array, no markdown fences."""

def extract_competencies(jd_text):
    api_key = os.environ.get("ANTHROPIC_API_KEY") or st.secrets.get("ANTHROPIC_API_KEY", "")
    if not api_key: return []
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-opus-4-8", max_tokens=300,
        messages=[{"role": "user", "content":
            f"Extract exactly 5 key skill competencies from this job description. "
            f"Return ONLY a JSON array of 5 short strings (3-5 words each), nothing else.\n\n{jd_text}"}]
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("` \n")
    return json.loads(raw)

def screen_cvs(jd, competencies, cvs, status_placeholder=None):
    api_key = os.environ.get("ANTHROPIC_API_KEY") or st.secrets.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        st.error("⚠️ ANTHROPIC_API_KEY is not set in Streamlit Secrets.")
        return None
    client = anthropic.Anthropic(api_key=api_key)
    if status_placeholder:
        status_placeholder.info(f"🔍 Analysing {len(cvs)} CVs with Claude… this takes 30–90 seconds.")
    try:
        msg = client.messages.create(
            model="claude-opus-4-8", max_tokens=8000,
            messages=[{"role": "user", "content": build_prompt(jd, competencies, cvs)}]
        )
        result = msg.content[0].text.strip()
    except Exception as e:
        if status_placeholder: status_placeholder.error(f"❌ Claude API error: {e}")
        else: st.error(f"❌ Claude API error: {e}")
        return None
    if status_placeholder: status_placeholder.empty()
    return result

# ── PDF export ────────────────────────────────────────────────────────────────

def generate_pdf(results, role_title=""):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    QB = colors.HexColor("#0075BC")
    G, R = colors.HexColor("#2E7D32"), colors.HexColor("#C62828")
    LG   = colors.HexColor("#F5F5F5")
    def sty(name, **kw): return ParagraphStyle(name, **kw)
    T  = sty("T",  fontSize=22, textColor=colors.HexColor("#1A1A2E"), spaceAfter=2*mm, fontName="Helvetica-Bold")
    ST = sty("ST", fontSize=11, textColor=colors.gray, spaceAfter=6*mm, fontName="Helvetica")
    SC = sty("SC", fontSize=13, textColor=QB, spaceBefore=6*mm, spaceAfter=2*mm, fontName="Helvetica-Bold")
    NM = sty("NM", fontSize=12, textColor=colors.HexColor("#1A1A2E"), fontName="Helvetica-Bold", spaceAfter=1*mm)
    BD = sty("BD", fontSize=9,  textColor=colors.HexColor("#333"), fontName="Helvetica", spaceAfter=2*mm, leading=13)
    LB = sty("LB", fontSize=9,  textColor=colors.gray, fontName="Helvetica-Bold", spaceAfter=1*mm)
    FT = sty("FT", fontSize=8,  textColor=colors.gray, fontName="Helvetica-Oblique")

    sl  = [r for r in results if r.get("shortlisted")]
    no  = [r for r in results if not r.get("shortlisted")]
    date_str = datetime.now().strftime("%d %B %Y")
    story = [Paragraph("CV Screening Report", T)]
    role_part = f"Role: {role_title} · " if role_title else ""
    story.append(Paragraph(f"{role_part}Generated: {date_str} · Quest Alliance", ST))
    story.append(HRFlowable(width="100%", thickness=2, color=QB, spaceAfter=6*mm))
    hdr = Table([["Total Screened", "Shortlisted", "Not Shortlisted"]], colWidths=[55*mm]*3)
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),10),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),3*mm),("BOTTOMPADDING",(0,0),(-1,-1),3*mm),
    ]))
    summ = Table([[str(len(results)), str(len(sl)), str(len(no))]], colWidths=[55*mm]*3)
    summ.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),11),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4*mm),("BOTTOMPADDING",(0,0),(-1,-1),4*mm),
        ("BOX",(0,0),(-1,-1),.5,colors.lightgrey),("INNERGRID",(0,0),(-1,-1),.5,colors.lightgrey),
    ]))
    story.extend([hdr, summ, Spacer(1,8*mm)])

    def cblock(r, is_sl):
        sc = r.get("overall", 0)
        sl_label = "SHORTLISTED" if is_sl else "NOT SHORTLISTED"
        sl_color = G if is_sl else R
        story.append(Paragraph(f"#{r.get('rank','')} · {r.get('name', r.get('filename',''))}", NM))
        story.append(Paragraph(
            f'<font color="{sl_color.hexval()}">{sl_label}</font>  ·  Score: <b>{sc}/100</b>  ·  '
            f'{r.get("role","")}  ·  {r.get("location","")}', BD))
        story.append(Paragraph(r.get("summary",""), BD))
        ct = Table([[Paragraph("Strengths", LB), Paragraph("Gaps", LB)],
                    [Paragraph(" · ".join(r.get("strengths",[])) or "—", BD),
                     Paragraph(" · ".join(r.get("gaps",[])) or "—", BD)]],
                   colWidths=[82*mm, 82*mm])
        ct.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),LG),("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),3*mm),("RIGHTPADDING",(0,0),(-1,-1),3*mm),
            ("TOPPADDING",(0,0),(-1,-1),2*mm),("BOTTOMPADDING",(0,0),(-1,-1),2*mm),
            ("BOX",(0,0),(-1,-1),.5,colors.lightgrey),("INNERGRID",(0,0),(-1,-1),.5,colors.lightgrey),
        ]))
        story.extend([ct, Spacer(1,5*mm),
                      HRFlowable(width="100%", thickness=.5, color=colors.lightgrey, spaceAfter=4*mm)])

    if sl:
        story.append(Paragraph(f"✓ Shortlisted ({len(sl)})", SC))
        story.append(HRFlowable(width="100%", thickness=1, color=G, spaceAfter=4*mm))
        for r in sl: cblock(r, True)
    if no:
        story.append(Paragraph(f"✗ Not Shortlisted ({len(no)})", SC))
        story.append(HRFlowable(width="100%", thickness=1, color=R, spaceAfter=4*mm))
        for r in no: cblock(r, False)
    story.append(Spacer(1,4*mm))
    story.append(Paragraph(
        "Generated by Quest CV Screener · Powered by Claude AI · "
        "Scores are AI-generated and should be used alongside human review.", FT))
    doc.build(story)
    return buf.getvalue()

# ── Excel export ─────────────────────────────────────────────────────────────

def generate_excel(results, role_title="", history=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    BLUE  = "FF0075BC"; GREEN = "FF1B6E2E"; RED = "FFC62828"
    ORG = "FFF7941D"; LGREY = "FFF4F7FA"; WHITE = "FFFFFFFF"; DARK = "FF1A1A2E"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=DARK)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="FFE4E9EF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_row = [f"CV Screening Report — {role_title}" if role_title else "CV Screening Report",
                 "", "", "", "", f"Generated: {datetime.now().strftime('%d %B %Y')}"]
    ws.append(title_row)
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=BLUE)
    ws.merge_cells("A1:E1")
    ws["F1"].font = Font(name="Calibri", size=10, color="FF5E6675")
    ws["F1"].alignment = Alignment(horizontal="right")
    ws.append([])

    columns = ["Rank","Name","Role","Years Exp","Location","Score (Reviewed)","Band","Status",
               "Review Status","Strengths","Gaps","Summary","Flag","Email","Phone","Filename"]
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border

    band_colors = {"strong": GREEN, "possible": ORG, "weak": RED}

    for r in results:
        sc = r.get("overall", 0)
        bname, _, _, _ = band(sc)
        sl = r.get("shortlisted")
        status = "Shortlisted" if sl is True else "Rejected" if sl is False else "Pending"
        ck = candidate_key(r)
        score_fb = st.session_state.score_feedback.get(ck, {})
        reviewed_score = score_fb.get("human_overall", sc) if score_fb else sc
        review_status  = ("Approved" if score_fb.get("approved") else
                          f"Calibrated ({sc}→{reviewed_score})") if score_fb else "Not reviewed"
        row = [r.get("rank",""), r.get("name", r.get("filename","")),
               r.get("role",""), r.get("years",""), r.get("location",""),
               reviewed_score, bname.capitalize(), status, review_status,
               "; ".join(r.get("strengths",[])), "; ".join(r.get("gaps",[])),
               r.get("summary",""), r.get("flag","") or "",
               r.get("email",""), r.get("phone",""), r.get("filename","")]
        ws.append(row)
        data_row = ws.max_row
        for col_idx in range(1, len(columns)+1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            if data_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LGREY)
        band_cell = ws.cell(row=data_row, column=7)
        fc = band_colors.get(bname, DARK)
        band_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        band_cell.fill = PatternFill("solid", fgColor=fc)
        st_cell = ws.cell(row=data_row, column=8)
        if sl is True:
            st_cell.font = Font(name="Calibri", bold=True, color=WHITE)
            st_cell.fill = PatternFill("solid", fgColor=GREEN)
        elif sl is False:
            st_cell.font = Font(name="Calibri", bold=True, color=WHITE)
            st_cell.fill = PatternFill("solid", fgColor=RED)
        rv_cell = ws.cell(row=data_row, column=9)
        if score_fb:
            rv_cell.font = Font(name="Calibri", bold=True,
                                color=GREEN if score_fb.get("approved") else BLUE, size=10)

    col_widths = [6, 22, 22, 10, 16, 14, 10, 12, 18, 40, 30, 50, 35, 24, 16, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    if history:
        wh = wb.create_sheet("Status History")
        wh.append(["Candidate", "Action", "Timestamp"])
        for col_idx in range(1, 4):
            cell = wh.cell(row=1, column=col_idx)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
        all_entries = sorted(
            [e for entries in history.values() for e in entries],
            key=lambda e: e["ts"], reverse=True
        )
        for e in all_entries:
            wh.append([e.get("name",""), e.get("action","").capitalize(), e.get("ts","")])
            row_idx = wh.max_row
            for col_idx in range(1, 4):
                wh.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center")
                wh.cell(row=row_idx, column=col_idx).border = border
        wh.column_dimensions["A"].width = 24
        wh.column_dimensions["B"].width = 16
        wh.column_dimensions["C"].width = 22
        wh.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── Session state ─────────────────────────────────────────────────────────────
for k, v in {"screen": "setup", "selected_idx": 0, "screening_results": None,
             "cvs": {}, "jd": "", "competencies": [], "role_title": "",
             "filter": "all", "search": "", "sort": "match",
             "jd_last_detected": "", "candidate_history": {},
             "score_feedback": {}, "feedback_examples": []}.items():
    if k not in st.session_state: st.session_state[k] = v

# ── Shared nav bar ────────────────────────────────────────────────────────────
screen = st.session_state.screen
sl_count = sum(1 for r in (st.session_state.screening_results or []) if r.get("shortlisted") is True)

components.html(f"""
<link href="https://fonts.googleapis.com/css2?family=Work+Sans:wght@500;600;700;800&display=swap" rel="stylesheet">
<div style="height:60px;background:#fff;border-bottom:1px solid #E4E9EF;display:flex;align-items:center;
justify-content:space-between;padding:0 28px;font-family:'Work Sans',sans-serif">
  <div style="display:flex;align-items:center;gap:11px">
    <div style="width:30px;height:30px;border-radius:9px;background:#0075BC;color:#fff;display:flex;
    align-items:center;justify-content:center;font-weight:800;font-size:16px">Q</div>
    <div>
      <div style="font-weight:700;font-size:15px;line-height:1.05;color:#222838">CV Screener</div>
      <div style="font:500 10px 'Work Sans';color:#9AA1AE">Quest Alliance · Enabling Self Learning</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:14px">
    {"" if screen == "setup" else f'<div style="font:600 13px Work Sans;color:#0075BC;background:#fff;border:1px solid #B8DCF0;border-radius:9px;padding:7px 13px">⤓ Export shortlist ({sl_count})</div>'}
    <div style="width:32px;height:32px;border-radius:50%;background:#F7941D;color:#fff;display:flex;
    align-items:center;justify-content:center;font-weight:700;font-size:12px">QA</div>
  </div>
</div>
""", height=62, scrolling=False)

# ═══════════════════════════════════════════════════════════════════════════════
# SETUP SCREEN
# ═══════════════════════════════════════════════════════════════════════════════
if screen == "setup":

    # ── Stepper ───────────────────────────────────────────────────────────────
    jd_done    = bool(st.session_state.jd)
    comps_done = bool(st.session_state.competencies)
    cvs_done   = bool(st.session_state.cvs)
    active_step = 4 if cvs_done else (3 if comps_done else (2 if jd_done else 1))

    def step_html(num, label, active):
        if active:
            circ = (f'<div style="width:26px;height:26px;border-radius:50%;background:#0075BC;'
                    f'color:#fff;font-weight:700;font-size:12px;display:flex;align-items:center;'
                    f'justify-content:center;flex:none">{num}</div>')
            txt  = f'<span style="font-weight:700;font-size:13px;color:#0075BC">{label}</span>'
        else:
            circ = (f'<div style="width:26px;height:26px;border-radius:50%;background:#fff;'
                    f'border:2px solid #CBD2DC;color:#9AA1AE;font-weight:700;font-size:12px;'
                    f'display:flex;align-items:center;justify-content:center;flex:none">{num}</div>')
            txt  = f'<span style="font-weight:600;font-size:13px;color:#9AA1AE">{label}</span>'
        return f'<div style="display:flex;align-items:center;gap:9px">{circ}{txt}</div>'

    steps = [("1","Role"), ("2","Criteria"), ("3","Upload CVs"), ("4","Review")]
    step_parts = []
    for i, (num, label) in enumerate(steps):
        step_parts.append(step_html(num, label, (i+1) <= active_step))
        if i < 3:
            grad = ("linear-gradient(90deg,#0075BC,#E4E9EF)" if i == 0 and active_step > 1
                    else ("#0075BC" if (i+1) < active_step else "#E4E9EF"))
            step_parts.append(f'<div style="flex:1;height:2px;background:{grad};margin:0 14px"></div>')

    auto_html(
        f'<div style="display:flex;align-items:center;padding:18px 0 14px;margin-bottom:4px">'
        f'{"".join(step_parts)}</div>'
    )

    # ── Page title ────────────────────────────────────────────────────────────
    auto_html("""
<div style="padding:4px 0 18px">
  <div style="font:800 28px 'Work Sans',sans-serif;letter-spacing:-.02em;color:#222838">
    Let's find your strongest candidates</div>
  <div style="font:500 15px 'Work Sans',sans-serif;color:#5E6675;margin-top:5px">
    Describe the role and what great looks like. We'll rank every CV against it.</div>
</div>
""")

    role_title = st.text_input("Role title", value=st.session_state.role_title,
                                placeholder="e.g. Placement Officer – Chennai")
    st.session_state.role_title = role_title

    # ── JD card ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        hc1, hc2 = st.columns([3, 1])
        with hc1:
            st.markdown('<div style="font-weight:700;font-size:16px;color:#222838;margin-bottom:6px">Job description</div>',
                        unsafe_allow_html=True)
        with hc2:
            jd_mode = st.radio("jd_mode", ["Paste text", "Upload file"], horizontal=True,
                                label_visibility="collapsed", key="jd_mode_radio")
        jd_input = ""
        if jd_mode == "Paste text":
            jd_input = st.text_area("jd_text", height=120,
                placeholder="Senior Program Associate — Employability. Lead facilitation of youth career-readiness programs…",
                label_visibility="collapsed", value=st.session_state.jd, key="jd_textarea")
            st.session_state.jd = jd_input
        else:
            jd_file = st.file_uploader("Upload JD", type=["pdf","docx","doc","txt"],
                                        key="jd_file_upload", label_visibility="collapsed")
            if jd_file:
                try:
                    jd_input = extract_file_text(jd_file)
                    st.session_state.jd = jd_input
                    st.success(f"✅ {jd_file.name}")
                except Exception as e:
                    st.error(str(e))
            else:
                jd_input = st.session_state.jd

    # ── Auto-detect competencies ───────────────────────────────────────────────
    current_jd = st.session_state.jd or jd_input
    if current_jd and current_jd != st.session_state.jd_last_detected and len(current_jd) > 80:
        st.session_state.jd_last_detected = current_jd
        with st.spinner("Auto-detecting skill competencies from JD…"):
            try:
                detected = extract_competencies(current_jd)
                if detected:
                    st.session_state.competencies = detected
                    st.rerun()
            except Exception:
                pass

    # ── Competencies card ─────────────────────────────────────────────────────
    with st.container(border=True):
        cc1, cc2 = st.columns([3, 1])
        with cc1:
            st.markdown('<div style="font-weight:700;font-size:16px;color:#222838">Skill competencies</div>',
                        unsafe_allow_html=True)
        with cc2:
            st.caption("Auto-detected · tap × to remove")
        comps = st.session_state.competencies
        if comps:
            pill_cols = st.columns(min(len(comps), 5))
            for i, (col, c) in enumerate(zip(pill_cols, comps[:5])):
                with col:
                    if st.button(f"{c} ×", key=f"rm_{i}", type="secondary"):
                        st.session_state.competencies = [x for x in st.session_state.competencies if x != c]
                        st.rerun()
            if len(comps) > 5:
                extra_cols = st.columns(min(len(comps)-5, 5))
                for i, (col, c) in enumerate(zip(extra_cols, comps[5:])):
                    with col:
                        if st.button(f"{c} ×", key=f"rm_{i+5}", type="secondary"):
                            st.session_state.competencies = [x for x in st.session_state.competencies if x != c]
                            st.rerun()
        nc1, nc2 = st.columns([5, 1])
        with nc1:
            new_comp = st.text_input("new_comp", placeholder="+ Add skill and press Add",
                                      label_visibility="collapsed", key="new_comp_input")
        with nc2:
            if st.button("Add", key="add_comp_btn", type="secondary"):
                if new_comp.strip() and new_comp.strip() not in st.session_state.competencies:
                    st.session_state.competencies.append(new_comp.strip())
                st.rerun()

    # ── CV upload card ────────────────────────────────────────────────────────
    n_cvs = len(st.session_state.cvs)
    with st.container(border=True):
        cv_hc1, cv_hc2 = st.columns([3, 1])
        with cv_hc1:
            auto_html(
                f'<div style="font-weight:700;font-size:16px;color:#222838">'
                f'Candidate CVs <span style="color:#0075BC">· {n_cvs} added</span></div>'
            )
        with cv_hc2:
            st.caption("PDF, DOCX or TXT")

        uploaded = st.file_uploader("Drop CVs here", type=["pdf","docx","doc","txt"],
                                     accept_multiple_files=True, label_visibility="visible")
        if uploaded:
            new_files = [f for f in uploaded if f.name not in st.session_state.cvs]
            if new_files:
                with st.spinner(f"Reading {len(new_files)} file(s)…"):
                    errors = []
                    for f in new_files:
                        try:
                            f.seek(0)
                            st.session_state.cvs[f.name] = extract_file_text(f)
                        except Exception as e:
                            errors.append(f"{f.name}: {e}")
                if errors:
                    st.warning("Some files could not be read:\n" + "\n".join(errors))
                st.rerun()

        names = list(st.session_state.cvs.keys())
        if names:
            # 2-col grid matching design
            cols = st.columns(2)
            for idx, nm in enumerate(names):
                ext = "pdf" if nm.lower().endswith(".pdf") else "doc"
                tag = "PDF" if ext == "pdf" else "DOC"
                bg  = "#FEF0DC" if ext == "pdf" else "#E4EDF7"
                fg  = "#C76A0A" if ext == "pdf" else "#3251A3"
                with cols[idx % 2]:
                    auto_html(
                        f'<div class="qs-file-chip">'
                        f'<span style="width:24px;height:24px;border-radius:6px;background:{bg};'
                        f'color:{fg};font-weight:700;font-size:9px;display:flex;align-items:center;'
                        f'justify-content:center;flex:none">{tag}</span>'
                        f'<span style="font-weight:600;font-size:12px;overflow:hidden;white-space:nowrap;'
                        f'text-overflow:ellipsis;flex:1;color:#222838">{nm}</span>'
                        f'<span style="color:#0075BC;font-size:13px">✓</span></div>'
                    )
            if st.button("Clear all CVs", key="clear_cvs", type="secondary"):
                st.session_state.cvs = {}
                st.rerun()

    # ── Footer / Screen button ────────────────────────────────────────────────
    jd_ready  = bool(st.session_state.jd or jd_input)
    cvs_ready = bool(st.session_state.cvs)
    can_screen = jd_ready and cvs_ready
    n_screen   = len(st.session_state.cvs)

    status_ph = st.empty()
    st.divider()

    fl, fr = st.columns([5, 2])
    with fl:
        if not jd_ready:
            st.caption("⚠️ Add a job description to continue")
        elif not cvs_ready:
            st.caption("⚠️ Upload at least one CV to continue")
        else:
            st.caption(f"Takes about 30–60 seconds for {n_screen} CV{'s' if n_screen != 1 else ''}")
    with fr:
        if st.button(f"Screen {n_screen} CVs →", type="primary",
                      disabled=not can_screen, key="screen_btn"):
            jd = st.session_state.jd or jd_input
            comps_str = "\n".join(st.session_state.competencies)
            raw = screen_cvs(jd, comps_str, st.session_state.cvs, status_placeholder=status_ph)
            if raw:
                raw = re.sub(r"^```[a-z]*\n?", "", raw.strip()).rstrip("` \n")
                try:
                    results = json.loads(raw)
                    for r in results:
                        if r.get("shortlisted"):
                            record_history(r, "auto-shortlisted")
                    st.session_state.screening_results = results
                    st.session_state.screen = "results"
                    st.rerun()
                except json.JSONDecodeError as e:
                    st.error(f"❌ Could not parse Claude's response as JSON: {e}")
                    st.code(raw[:3000])
            elif raw is None:
                pass  # error already shown in screen_cvs

# ═══════════════════════════════════════════════════════════════════════════════
# RESULTS SCREEN
# ═══════════════════════════════════════════════════════════════════════════════
elif screen == "results":
    results_data  = st.session_state.screening_results or []
    strong_list   = [r for r in results_data if r.get("overall", 0) >= 85]
    possible_list = [r for r in results_data if 65 <= r.get("overall", 0) < 85]
    weak_list     = [r for r in results_data if r.get("overall", 0) < 65]
    role_label    = st.session_state.role_title or "Screening Results"
    filt          = st.session_state.filter

    # ── Header: breadcrumb + title + band stat pills ──────────────────────────
    auto_html(f"""
<div style="padding:24px 0 16px;border-bottom:1px solid #E4E9EF;margin-bottom:18px">
  <div style="display:flex;align-items:flex-end;justify-content:space-between;gap:16px;flex-wrap:wrap">
    <div>
      <div style="font:500 13px 'Work Sans',sans-serif;color:#5E6675;margin-bottom:5px">
        ← {role_label}</div>
      <div style="font:800 24px 'Work Sans',sans-serif;letter-spacing:-.02em;color:#222838">
        {len(results_data)} candidates screened, ranked</div>
    </div>
    <div style="display:flex;gap:10px;flex:none">
      <div style="text-align:center;background:#E3F1FA;border-radius:11px;padding:9px 16px;min-width:72px">
        <div style="font:800 20px 'Work Sans';color:#005A91">{len(strong_list)}</div>
        <div style="font:600 11px 'Work Sans';color:#005A91">Strong</div>
      </div>
      <div style="text-align:center;background:#FEF0DC;border-radius:11px;padding:9px 16px;min-width:72px">
        <div style="font:800 20px 'Work Sans';color:#C76A0A">{len(possible_list)}</div>
        <div style="font:600 11px 'Work Sans';color:#C76A0A">Possible</div>
      </div>
      <div style="text-align:center;background:#FDE7DE;border-radius:11px;padding:9px 16px;min-width:72px">
        <div style="font:800 20px 'Work Sans';color:#C23A18">{len(weak_list)}</div>
        <div style="font:600 11px 'Work Sans';color:#C23A18">Weak</div>
      </div>
    </div>
  </div>
</div>
""")

    # ── Toolbar: search + filter buttons + sort ───────────────────────────────
    tb1, tb2 = st.columns([5, 3])
    with tb1:
        search = st.text_input("search", placeholder="🔍  Search candidates…",
                                label_visibility="collapsed", key="search_input")
    with tb2:
        sort = st.selectbox("sort", ["Best match", "Name A–Z", "Experience"],
                             label_visibility="collapsed", key="sort_sel")

    fb1, fb2, fb3, fb4 = st.columns(4)
    with fb1:
        if st.button("All", key="f_all",
                     type="primary" if filt == "all" else "secondary"):
            st.session_state.filter = "all"; st.rerun()
    with fb2:
        if st.button(f"Strong · {len(strong_list)}", key="fband_strong",
                     type="primary" if filt == "strong" else "secondary"):
            st.session_state.filter = "strong" if filt != "strong" else "all"; st.rerun()
    with fb3:
        if st.button(f"Possible · {len(possible_list)}", key="fband_possible",
                     type="primary" if filt == "possible" else "secondary"):
            st.session_state.filter = "possible" if filt != "possible" else "all"; st.rerun()
    with fb4:
        if st.button(f"Weak · {len(weak_list)}", key="fband_weak",
                     type="primary" if filt == "weak" else "secondary"):
            st.session_state.filter = "weak" if filt != "weak" else "all"; st.rerun()

    # ── Filter & sort ─────────────────────────────────────────────────────────
    filtered = results_data
    if filt == "strong":     filtered = strong_list
    elif filt == "possible": filtered = possible_list
    elif filt == "weak":     filtered = weak_list
    if search:
        filtered = [r for r in filtered
                    if search.lower() in r.get("name","").lower()
                    or search.lower() in r.get("filename","").lower()]
    if "Name" in sort:  filtered = sorted(filtered, key=lambda r: r.get("name",""))
    elif "Exp" in sort: filtered = sorted(filtered, key=lambda r: -r.get("years", 0))

    if not filtered:
        auto_html('<div style="text-align:center;font:600 14px Work Sans;color:#9AA1AE;padding:40px 0">No candidates match these filters.</div>')

    # ── Candidate rows ────────────────────────────────────────────────────────
    for i, r in enumerate(filtered):
        sc         = r.get("overall", 0)
        _, bbg, bcolor, border_color = band(sc)
        av_color   = avatar_color(r.get("name", r.get("filename","")))
        name       = r.get("name", r.get("filename",""))
        inits      = initials(name)
        strengths  = r.get("strengths", [])
        tags       = strengths[:2]
        extra_tags = len(strengths) - 2

        tags_html = "".join(
            f'<span style="background:#E3F1FA;color:#005A91;border-radius:6px;'
            f'padding:4px 9px;font-weight:600;font-size:11px;white-space:nowrap">{t}</span>'
            for t in tags)
        if extra_tags > 0:
            tags_html += (f'<span style="background:#EDF0F4;color:#5E6675;border-radius:6px;'
                          f'padding:4px 9px;font-weight:600;font-size:11px">+{extra_tags}</span>')

        is_sl = r.get("shortlisted") is True
        is_rj = r.get("shortlisted") is False
        ckey_r = candidate_key(r)
        fb_r   = st.session_state.score_feedback.get(ckey_r)

        status_badges = ""
        if is_sl:
            status_badges += '<span style="background:#E3F1FA;color:#005A91;border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700">★ Shortlisted</span>'
        elif is_rj:
            status_badges += '<span style="background:#FDE7DE;color:#C23A18;border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700">✕ Rejected</span>'
        if fb_r:
            cal_label = "✓ Approved" if fb_r.get("approved") else "✎ Calibrated"
            status_badges += f'<span style="background:#F0FDF4;color:#1B6E2E;border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700;margin-left:4px">{cal_label}</span>'

        global_idx = next((j for j, x in enumerate(results_data)
                           if candidate_key(x) == ckey_r), i)

        col_main, col_actions = st.columns([8, 2])
        with col_main:
            row_opacity = "0.55" if is_rj else "1"
            auto_html(f"""
<div class="qs-cand-card" style="display:flex;align-items:center;gap:16px;background:#fff;
border:1px solid #E4E9EF;border-left:4px solid {border_color};border-radius:12px;
padding:13px 18px;margin-bottom:2px;opacity:{row_opacity}">
  <div style="font:800 15px 'Work Sans';color:#C2C8D2;width:20px;flex:none;text-align:left">
    {r.get('rank', i+1)}</div>
  <div style="width:42px;height:42px;border-radius:50%;background:{av_color};color:#fff;
  display:flex;align-items:center;justify-content:center;font-weight:700;font-size:15px;flex:none">
    {inits}</div>
  <div style="flex:1;min-width:0">
    <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
      <span style="font-weight:700;font-size:15px;color:#222838">{name}</span>
      {status_badges}
    </div>
    <div style="font:500 13px 'Work Sans';color:#5E6675;margin-top:2px">
      {r.get('role','')} · {r.get('years','')} yrs · {r.get('location','')}</div>
    <div style="display:flex;gap:6px;margin-top:7px;flex-wrap:wrap">{tags_html}</div>
  </div>
  {score_ring(sc, 54)}
</div>
""")

        with col_actions:
            a1, a2, a3 = st.columns(3)
            with a1:
                if st.button("→", key=f"view_{i}", help="View detail", type="secondary"):
                    st.session_state.selected_idx = global_idx
                    st.session_state.screen = "detail"
                    st.rerun()
            with a2:
                sl_title = "undo-shortlist" if is_sl else "Shortlist candidate"
                if st.button("★", key=f"sl_{i}", help=sl_title, type="secondary"):
                    if is_sl:
                        st.session_state.screening_results[global_idx]["shortlisted"] = None
                        record_history(r, "un-shortlisted")
                    else:
                        st.session_state.screening_results[global_idx]["shortlisted"] = True
                        record_history(r, "shortlisted")
                    st.rerun()
            with a3:
                rj_title = "undo-reject" if is_rj else "Reject candidate"
                if st.button("✕", key=f"rj_{i}", help=rj_title, type="secondary"):
                    if is_rj:
                        st.session_state.screening_results[global_idx]["shortlisted"] = None
                        record_history(r, "un-rejected")
                    else:
                        st.session_state.screening_results[global_idx]["shortlisted"] = False
                        record_history(r, "rejected")
                    st.rerun()

    st.divider()

    # ── Candidate history ─────────────────────────────────────────────────────
    history = st.session_state.candidate_history
    if history:
        with st.expander(f"📋 Candidate status history ({sum(len(v) for v in history.values())} actions)"):
            all_entries = sorted(
                [e for entries in history.values() for e in entries],
                key=lambda e: e["ts"], reverse=True
            )
            for e in all_entries:
                action = e["action"]
                icon, color = (("★", "#005A91") if action == "shortlisted"
                               else ("✕", "#C23A18") if action == "rejected"
                               else ("·", "#5E6675"))
                auto_html(
                    f'<div style="display:flex;align-items:center;gap:10px;padding:8px 0;'
                    f'border-bottom:1px solid #F0F2F5">'
                    f'<span style="color:{color};font-weight:700;font-size:14px;width:16px">{icon}</span>'
                    f'<span style="font-weight:600;font-size:13px;color:#222838;flex:1">{e["name"]}</span>'
                    f'<span style="font-size:12px;color:#5E6675">{action.capitalize()}</span>'
                    f'<span style="font-size:11px;color:#9AA1AE;margin-left:12px">{e["ts"]}</span>'
                    f'</div>'
                )

    c1, c2, c3, _ = st.columns([2, 2, 2, 2])
    with c1:
        if st.button("← New screening", key="back_to_setup"):
            st.session_state.screen = "setup"
            st.session_state.screening_results = None
            st.rerun()
    with c2:
        pdf = generate_pdf(results_data, st.session_state.role_title)
        st.download_button("📄 Download PDF", data=pdf,
            file_name=f"CV_Screening_{datetime.now().strftime('%Y-%m-%d')}.pdf",
            mime="application/pdf", type="primary")
    with c3:
        xlsx = generate_excel(results_data, st.session_state.role_title,
                              st.session_state.candidate_history)
        st.download_button("📊 Download Excel", data=xlsx,
            file_name=f"CV_Screening_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary")

# ═══════════════════════════════════════════════════════════════════════════════
# DETAIL SCREEN
# ═══════════════════════════════════════════════════════════════════════════════
elif screen == "detail":
    results_data = st.session_state.screening_results or []
    idx = st.session_state.selected_idx
    if not results_data or idx >= len(results_data):
        st.session_state.screen = "results"; st.rerun()

    r        = results_data[idx]
    sc       = r.get("overall", 0)
    name     = r.get("name", r.get("filename", ""))
    inits    = initials(name)
    av_color = avatar_color(name)
    rc       = ring_color(sc)
    bname, bbg, bcolor, _ = band(sc)
    band_label = {"strong": "Strong match", "possible": "Possible", "weak": "Weak match"}[bname]
    scores   = r.get("scores", [])
    labels   = r.get("competency_labels", [f"Competency {i+1}" for i in range(len(scores))])

    # ── Nav ───────────────────────────────────────────────────────────────────
    nav1, _, nav3 = st.columns([3, 4, 3])
    with nav1:
        if st.button("← Back to candidates", key="back_results"):
            st.session_state.screen = "results"; st.rerun()
    with nav3:
        pc1, pc2 = st.columns(2)
        with pc1:
            if idx > 0 and st.button("↑ Prev", key="prev_btn"):
                st.session_state.selected_idx -= 1; st.rerun()
        with pc2:
            if idx < len(results_data)-1 and st.button("Next ↓", key="next_btn"):
                st.session_state.selected_idx += 1; st.rerun()

    # ── Identity header ───────────────────────────────────────────────────────
    email_part = f'<span>✉ {r["email"]}</span>' if r.get("email") else ""
    phone_part = f'<span>📞 {r["phone"]}</span>' if r.get("phone") else ""
    contact_row = f'<div style="display:flex;gap:16px;margin-top:9px;font:500 13px Work Sans;color:#3A4150;flex-wrap:wrap">{email_part}{phone_part}</div>' if (email_part or phone_part) else ""

    detail_ring = score_ring(sc, 78)

    auto_html(f"""
<div style="background:#fff;border:1px solid #E4E9EF;border-radius:14px;
display:flex;align-items:center;gap:22px;padding:22px 24px;margin:16px 0 0">
  <div style="width:74px;height:74px;border-radius:50%;background:{av_color};color:#fff;
  display:flex;align-items:center;justify-content:center;font-weight:700;font-size:27px;flex:none">{inits}</div>
  <div style="flex:1;min-width:0">
    <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
      <span style="font:800 24px 'Work Sans';letter-spacing:-.02em;color:#222838">{name}</span>
      <span style="background:{bbg};color:{bcolor};border-radius:999px;padding:4px 11px;
      font-weight:700;font-size:12px">{band_label}</span>
    </div>
    <div style="font:500 14px 'Work Sans';color:#5E6675;margin-top:4px">
      {r.get('role','')} · {r.get('years','')} yrs experience · {r.get('location','')}</div>
    {contact_row}
  </div>
  {detail_ring}
</div>
""")

    # ── Action bar ────────────────────────────────────────────────────────────
    auto_html('<div style="background:#F4F7FA;border-radius:12px;padding:12px 0;margin:12px 0 4px;'
              'display:flex;gap:10px;flex-wrap:wrap"></div>')  # visual spacer hint
    is_sl_detail = r.get("shortlisted") is True
    is_rj_detail = r.get("shortlisted") is False
    ac1, ac2, ac3, _ = st.columns([2, 2, 2, 4])
    with ac1:
        sl_title_d = "undo-shortlist" if is_sl_detail else "Shortlist candidate"
        sl_label   = "★ Shortlisted" if is_sl_detail else "★ Shortlist"
        if st.button(sl_label, key="sl_btn", help=sl_title_d,
                     type="primary" if not is_sl_detail else "secondary"):
            if is_sl_detail:
                st.session_state.screening_results[idx]["shortlisted"] = None
                record_history(r, "un-shortlisted")
            else:
                st.session_state.screening_results[idx]["shortlisted"] = True
                record_history(r, "shortlisted")
            st.rerun()
    with ac2:
        rj_title_d = "undo-reject" if is_rj_detail else "Reject candidate"
        rj_label   = "✕ Rejected" if is_rj_detail else "✕ Reject"
        if st.button(rj_label, key="rj_btn", help=rj_title_d, type="secondary"):
            if is_rj_detail:
                st.session_state.screening_results[idx]["shortlisted"] = None
                record_history(r, "un-rejected")
            else:
                st.session_state.screening_results[idx]["shortlisted"] = False
                record_history(r, "rejected")
            st.rerun()
    with ac3:
        pdf_single = generate_pdf([r], st.session_state.role_title)
        st.download_button("⤓ Export PDF", data=pdf_single,
            file_name=f"{name.replace(' ','_')}_Report.pdf",
            mime="application/pdf", type="secondary")

    # Status history for this candidate
    ckey = candidate_key(r)
    chistory = st.session_state.candidate_history.get(ckey, [])
    if chistory:
        st.caption("Status history: " + "  ·  ".join(
            f"{e['action'].capitalize()} at {e['ts']}" for e in reversed(chistory)))

    st.divider()

    left_col, right_col = st.columns([1.2, 1])

    with left_col:
        # ── Score calibration panel ───────────────────────────────────────────
        ckey_d    = candidate_key(r)
        existing_fb = st.session_state.score_feedback.get(ckey_d)

        if existing_fb:
            approved_txt   = "✓ Scores approved" if existing_fb.get("approved") else "✎ Scores calibrated by reviewer"
            approved_color = "#1B6E2E" if existing_fb.get("approved") else "#0075BC"
            auto_html(f"""
<div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:10px;
padding:10px 14px;margin-bottom:12px">
  <div style="font-weight:700;font-size:12px;color:{approved_color}">{approved_txt}</div>
  <div style="font:500 11px Work Sans;color:#5E6675;margin-top:3px;line-height:1.5">
    Overall: AI={existing_fb['ai_overall']} → Reviewer={existing_fb['human_overall']}
    {"  ·  " + existing_fb['reason'] if existing_fb.get('reason') else ""}
    · {existing_fb.get('ts','')}
  </div>
</div>""")

        with st.expander("🎯 Review & calibrate AI scores", expanded=not bool(existing_fb)):
            st.caption("Adjust scores if the AI got something wrong — corrections calibrate future screenings.")
            prior_overall = existing_fb["human_overall"] if existing_fb else sc
            prior_comp    = existing_fb["human_scores"]  if existing_fb else list(scores)

            new_overall = st.slider("Overall match score", 0, 100, prior_overall,
                                    key=f"cal_overall_{idx}")
            new_comp_scores = []
            for j, (label, _) in enumerate(zip(labels, prior_comp)):
                default_val = prior_comp[j] if j < len(prior_comp) else (scores[j] if j < len(scores) else 50)
                new_val = st.slider(label, 0, 100, int(default_val), key=f"cal_comp_{idx}_{j}")
                new_comp_scores.append(new_val)

            cal_reason = st.text_area(
                "Reason for adjustment (optional)",
                value=existing_fb.get("reason","") if existing_fb else "",
                placeholder="e.g. Strong field experience not reflected in CV text…",
                height=72, key=f"cal_reason_{idx}")

            cal1, cal2 = st.columns(2)
            with cal1:
                if st.button("✓ Approve as-is", key=f"cal_approve_{idx}", type="secondary"):
                    record_feedback(r, sc, list(scores), labels, sc, list(scores), "", approved=True)
                    st.session_state.screening_results[idx]["overall"] = sc
                    st.success("Scores approved."); st.rerun()
            with cal2:
                if st.button("💾 Save adjustments", key=f"cal_save_{idx}", type="primary"):
                    record_feedback(r, sc, list(scores), labels,
                                    new_overall, new_comp_scores, cal_reason.strip(), approved=False)
                    st.session_state.screening_results[idx]["overall"] = new_overall
                    st.session_state.screening_results[idx]["scores"]  = new_comp_scores
                    st.success("Saved — future screenings will learn from this."); st.rerun()

        # ── Score bars ────────────────────────────────────────────────────────
        auto_html('<div style="font:700 16px Work Sans;color:#222838;margin:16px 0 4px">Match breakdown</div>')
        auto_html('<div style="font:500 13px Work Sans;color:#5E6675;margin-bottom:14px">Scored against your 5 competencies</div>')

        display_scores  = existing_fb["human_scores"]  if existing_fb and not existing_fb.get("approved") else scores
        display_overall = existing_fb["human_overall"] if existing_fb else sc

        for label, val in zip(labels, display_scores):
            bar_color = "#0075BC" if val >= 80 else "#F7941D" if val >= 60 else "#F15A29"
            val_color = "#005A91" if val >= 80 else "#C76A0A" if val >= 60 else "#C23A18"
            auto_html(f"""
<div style="margin-bottom:16px">
  <div style="display:flex;justify-content:space-between;font-weight:600;font-size:14px;
  margin-bottom:7px;color:#222838">
    <span>{label}</span><span style="color:{val_color}">{val}</span>
  </div>
  <div style="height:9px;background:#E4E9EF;border-radius:999px;overflow:hidden">
    <div style="height:100%;width:{val}%;background:{bar_color};border-radius:999px"></div>
  </div>
</div>""")

        if r.get("flag"):
            auto_html(f"""
<div style="background:#FEF0DC;border-radius:11px;padding:14px 16px;
display:flex;gap:11px;margin-top:6px;margin-bottom:6px">
  <div style="font-size:16px;flex:none">⚠</div>
  <div>
    <div style="font-weight:700;font-size:13px;color:#B0640C">One thing to verify</div>
    <div style="font:500 13px/1.5 Work Sans;color:#9A6410;margin-top:2px">{r['flag']}</div>
  </div>
</div>""")

        auto_html('<div style="font:700 15px Work Sans;color:#222838;margin:16px 0 6px">Gaps / Concerns</div>')
        for g in r.get("gaps", []):
            auto_html(f'<div style="font:500 13px/1.6 Work Sans;color:#3A4150;padding:4px 0">· {g}</div>')

    with right_col:
        # ── AI summary ────────────────────────────────────────────────────────
        auto_html(f"""
<div style="background:#0075BC;border-radius:12px;padding:16px 18px;color:#fff;margin-bottom:18px">
  <div style="display:flex;align-items:center;gap:8px;font-weight:700;font-size:13px;margin-bottom:8px">
    <span style="width:18px;height:18px;border-radius:5px;background:rgba(255,255,255,.22);
    display:flex;align-items:center;justify-content:center;font-size:11px">✦</span>
    AI summary
  </div>
  <div style="font:500 13.5px/1.6 Work Sans;color:rgba(255,255,255,.92)">{r.get('summary','')}</div>
</div>""")

        # ── Evidence cards ────────────────────────────────────────────────────
        auto_html('<div style="font:700 15px Work Sans;color:#222838;margin-bottom:12px">Evidence from CV</div>')
        for ev in r.get("evidence", []):
            auto_html(f"""
<div style="background:#fff;border:1px solid #E4E9EF;border-radius:10px;
padding:12px 14px;margin-bottom:10px">
  <div style="font:600 11px 'JetBrains Mono',monospace;color:#0075BC;margin-bottom:4px">
    {ev.get('label','')}</div>
  <div style="font:500 13px/1.5 Work Sans;color:#3A4150">{ev.get('text','')}</div>
</div>""")
