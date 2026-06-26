import os, io, re, json
from datetime import datetime
from typing import List, Optional

import anthropic
import pdfplumber
from docx import Document
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from fastapi import Cookie, Depends, FastAPI, UploadFile, File, Form, HTTPException, Response as FResponse
from fastapi.responses import Response, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import auth as _auth

app = FastAPI(title="CV Screener – Quest Alliance")

# ── Startup ────────────────────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    _auth.init_db()

# ── Auth routes ────────────────────────────────────────────────────────────────

@app.post("/api/auth/login")
async def login(
    username: str = Form(...),
    password: str = Form(...),
    response: FResponse = None,
):
    user = _auth.get_user_by_credentials(username, password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = _auth.create_token(user["id"], user["username"], user["role"])
    resp = JSONResponse({"username": user["username"], "role": user["role"], "email": user["email"]})
    resp.set_cookie(
        _auth.COOKIE, token,
        httponly=True, samesite="lax", secure=False,  # set secure=True behind HTTPS
        max_age=_auth.TOKEN_TTL * 3600,
    )
    return resp

@app.post("/api/auth/logout")
async def logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(_auth.COOKIE)
    return resp

@app.get("/api/auth/me")
async def me(qs_token: Optional[str] = Cookie(default=None)):
    if not qs_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = _auth.get_current_user(qs_token)
    return {"username": user["username"], "role": user["role"], "email": user["email"]}

# ── Admin routes ───────────────────────────────────────────────────────────────

class NewUser(BaseModel):
    username: str
    email: str
    password: str
    role: str = "recruiter"

class UpdateUser(BaseModel):
    role: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None

@app.get("/api/admin/users")
async def admin_list_users(qs_token: Optional[str] = Cookie(default=None)):
    _auth.require_admin(qs_token=qs_token)
    return _auth.list_users()

@app.post("/api/admin/users")
async def admin_create_user(body: NewUser, qs_token: Optional[str] = Cookie(default=None)):
    _auth.require_admin(qs_token=qs_token)
    return _auth.create_user(body.username, body.email, body.password, body.role)

@app.patch("/api/admin/users/{uid}")
async def admin_update_user(uid: int, body: UpdateUser, qs_token: Optional[str] = Cookie(default=None)):
    _auth.require_admin(qs_token=qs_token)
    return _auth.update_user(uid, body.role, body.active, body.password)

@app.delete("/api/admin/users/{uid}")
async def admin_delete_user(uid: int, qs_token: Optional[str] = Cookie(default=None)):
    admin = _auth.require_admin(qs_token=qs_token)
    _auth.delete_user(uid, admin["id"])
    return {"ok": True}

# ── Calibration notes routes ───────────────────────────────────────────────────

class CalibNote(BaseModel):
    note: str

@app.get("/api/calibration")
async def get_calibration(qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    return _auth.get_calibration_notes()

@app.post("/api/calibration")
async def add_calibration(body: CalibNote, qs_token: Optional[str] = Cookie(default=None)):
    user = _auth.get_current_user(qs_token)
    return _auth.add_calibration_note(body.note, user["username"])

@app.delete("/api/calibration/{note_id}")
async def delete_calibration(note_id: int, qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    _auth.delete_calibration_note(note_id)
    return {"ok": True}

# ── File parsing ───────────────────────────────────────────────────────────────

def extract_text_from_pdf(b: bytes) -> str:
    with pdfplumber.open(io.BytesIO(b)) as p:
        return "\n".join(pg.extract_text() or "" for pg in p.pages)

def extract_text_from_docx(b: bytes) -> str:
    return "\n".join(para.text for para in Document(io.BytesIO(b)).paragraphs)

def parse_bytes(b: bytes, name: str) -> str:
    n = name.lower()
    if n.endswith(".pdf"):            return extract_text_from_pdf(b)
    if n.endswith((".docx", ".doc")): return extract_text_from_docx(b)
    return b.decode("utf-8", errors="replace")

# ── Helpers ────────────────────────────────────────────────────────────────────

def candidate_key(r: dict) -> str:
    return r.get("filename") or r.get("name") or str(r.get("rank", ""))

def band(score: int):
    if score >= 85: return "strong",   "#E3F1FA", "#005A91", "#0075BC"
    if score >= 65: return "possible", "#FEF0DC", "#C76A0A", "#F7941D"
    return "weak", "#FDE7DE", "#C23A18", "#F15A29"

def get_api_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not set on server.")
    return key

# ── Prompt ─────────────────────────────────────────────────────────────────────

def build_prompt(jd: str, competencies: str, cvs: dict, calibration: list,
                 past_examples: list = None) -> str:
    cv_block = "".join(
        f"\n---\nCV #{i} – {name}\n{text}\n"
        for i, (name, text) in enumerate(cvs.items(), 1)
    )
    calibration_block = ""
    if calibration:
        calibration_block = (
            "\n## Reviewer Calibration Notes\n"
            "Use these organisational preferences when scoring.\n"
            + "\n".join(f"- {ex}" for ex in calibration[-10:]) + "\n"
        )
    examples_block = ""
    if past_examples:
        lines = []
        for ex in past_examples[:15]:
            dec  = ex.get("final_decision", "").upper()
            score = ex.get("ai_score", "?")
            name  = ex.get("candidate_name", "Candidate")
            summ  = ex.get("summary", "")
            note  = ex.get("recruiter_note", "")
            line  = f"- {name} | AI score {score} → {dec}"
            if summ: line += f" | {summ[:120]}"
            if note: line += f" | Recruiter note: {note}"
            lines.append(line)
        examples_block = (
            "\n## Past Hiring Decisions for This Role (learn from these)\n"
            "These are real recruiter decisions for the same role. Calibrate your scoring "
            "so that candidates similar to SHORTLISTED examples score ≥65 and candidates "
            "similar to REJECTED examples score <65.\n"
            + "\n".join(lines) + "\n"
        )
    return f"""You are an expert HR screener for Quest Alliance, an NGO focused on youth skilling in India.

## Job Description
{jd}

## Required Skill Competencies
{competencies or "(derive from JD)"}
{calibration_block}{examples_block}
## Candidate CVs
{cv_block}

## Task
Review each candidate carefully and return a JSON array sorted best-to-worst. Each object MUST have:
- "rank": integer from 1
- "filename": CV filename exactly as given
- "name": candidate's full name (extract from CV)
- "role": their current/most recent role title
- "years": integer years of total experience
- "location": city/state from CV
- "email": email address if present, else ""
- "phone": phone if present, else ""
- "overall": integer 0-100 match score
- "scores": array of 5 integers (one per competency, same order as competencies)
- "competency_labels": array of 5 strings (the competency names scored)
- "shortlisted": true if overall >= 65
- "summary": 2-3 sentence AI summary of the candidate's fit
- "strengths": array of 2-4 short strength strings
- "gaps": array of 1-2 gap strings
- "flag": a one-sentence thing to verify, or null
- "evidence": array of 2-3 objects with "label" (competency name, uppercase) and "text" (direct quote or paraphrase from CV)

Return ONLY the JSON array, no markdown fences."""

def extract_json(raw: str) -> list:
    cleaned = re.sub(r"^```[a-z]*\n?", "", raw.strip()).rstrip("` \n")
    if not cleaned.strip():
        cleaned = raw.strip()
    m = re.search(r'\[[\s\S]*\]', cleaned)
    if m:
        cleaned = m.group(0)
    return json.loads(cleaned)

# ── PDF export ─────────────────────────────────────────────────────────────────

def generate_pdf(results: list, role_title: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    QB = colors.HexColor("#0075BC")
    G, R = colors.HexColor("#2E7D32"), colors.HexColor("#C62828")
    LG   = colors.HexColor("#F5F5F5")
    def sty(name, **kw): return ParagraphStyle(name, **kw)
    T  = sty("T",  fontSize=22, textColor=colors.HexColor("#1A1A2E"), spaceAfter=2*mm, fontName="Helvetica-Bold")
    ST = sty("ST", fontSize=11, textColor=colors.gray, spaceAfter=6*mm, fontName="Helvetica")
    SC = sty("SC", fontSize=13, textColor=QB, spaceBefore=6*mm, spaceAfter=2*mm, fontName="Helvetica-Bold")
    NM = sty("NM", fontSize=12, textColor=colors.HexColor("#1A1A2E"), fontName="Helvetica-Bold", spaceAfter=1*mm)
    BD = sty("BD", fontSize=9,  textColor=colors.HexColor("#333"), fontName="Helvetica", spaceAfter=2*mm, leading=13)
    LB = sty("LB", fontSize=9,  textColor=colors.gray, fontName="Helvetica-Bold", spaceAfter=1*mm)
    FT = sty("FT", fontSize=8,  textColor=colors.gray, fontName="Helvetica-Oblique")

    sl  = [r for r in results if r.get("shortlisted")]
    no  = [r for r in results if not r.get("shortlisted")]
    date_str = datetime.now().strftime("%d %B %Y")
    story = [Paragraph("CV Screening Report", T)]
    role_part = f"Role: {role_title} · " if role_title else ""
    story.append(Paragraph(f"{role_part}Generated: {date_str} · Quest Alliance", ST))
    story.append(HRFlowable(width="100%", thickness=2, color=QB, spaceAfter=6*mm))
    hdr = Table([["Total Screened", "Shortlisted", "Not Shortlisted"]], colWidths=[55*mm]*3)
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),10),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),3*mm),("BOTTOMPADDING",(0,0),(-1,-1),3*mm),
    ]))
    summ = Table([[str(len(results)), str(len(sl)), str(len(no))]], colWidths=[55*mm]*3)
    summ.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),11),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4*mm),("BOTTOMPADDING",(0,0),(-1,-1),4*mm),
        ("BOX",(0,0),(-1,-1),.5,colors.lightgrey),("INNERGRID",(0,0),(-1,-1),.5,colors.lightgrey),
    ]))
    story.extend([hdr, summ, Spacer(1,8*mm)])

    def cblock(r, is_sl):
        sc = r.get("overall", 0)
        sl_label = "SHORTLISTED" if is_sl else "NOT SHORTLISTED"
        sl_color = G if is_sl else R
        story.append(Paragraph(f"#{r.get('rank','')} · {r.get('name', r.get('filename',''))}", NM))
        story.append(Paragraph(
            f'<font color="{sl_color.hexval()}">{sl_label}</font>  ·  Score: <b>{sc}/100</b>  ·  '
            f'{r.get("role","")}  ·  {r.get("location","")}', BD))
        story.append(Paragraph(r.get("summary",""), BD))
        ct = Table([[Paragraph("Strengths", LB), Paragraph("Gaps", LB)],
                    [Paragraph(" · ".join(r.get("strengths",[])) or "—", BD),
                     Paragraph(" · ".join(r.get("gaps",[])) or "—", BD)]],
                   colWidths=[82*mm, 82*mm])
        ct.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),LG),("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),3*mm),("RIGHTPADDING",(0,0),(-1,-1),3*mm),
            ("TOPPADDING",(0,0),(-1,-1),2*mm),("BOTTOMPADDING",(0,0),(-1,-1),2*mm),
            ("BOX",(0,0),(-1,-1),.5,colors.lightgrey),("INNERGRID",(0,0),(-1,-1),.5,colors.lightgrey),
        ]))
        story.extend([ct, Spacer(1,5*mm),
                      HRFlowable(width="100%", thickness=.5, color=colors.lightgrey, spaceAfter=4*mm)])

    if sl:
        story.append(Paragraph(f"✓ Shortlisted ({len(sl)})", SC))
        story.append(HRFlowable(width="100%", thickness=1, color=G, spaceAfter=4*mm))
        for r in sl: cblock(r, True)
    if no:
        story.append(Paragraph(f"✗ Not Shortlisted ({len(no)})", SC))
        story.append(HRFlowable(width="100%", thickness=1, color=R, spaceAfter=4*mm))
        for r in no: cblock(r, False)
    story.append(Spacer(1,4*mm))
    story.append(Paragraph(
        "Generated by Quest CV Screener · Powered by Claude AI · "
        "Scores are AI-generated and should be used alongside human review.", FT))
    doc.build(story)
    return buf.getvalue()

# ── Excel export ───────────────────────────────────────────────────────────────

def generate_excel(results: list, role_title: str = "", history: dict = None, score_feedback: dict = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    score_feedback = score_feedback or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    BLUE  = "FF0075BC"; GREEN = "FF1B6E2E"; RED = "FFC62828"
    ORG = "FFF7941D"; LGREY = "FFF4F7FA"; WHITE = "FFFFFFFF"; DARK = "FF1A1A2E"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=DARK)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="FFE4E9EF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_row = [f"CV Screening Report — {role_title}" if role_title else "CV Screening Report",
                 "", "", "", "", f"Generated: {datetime.now().strftime('%d %B %Y')}"]
    ws.append(title_row)
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=BLUE)
    ws.merge_cells("A1:E1")
    ws["F1"].font = Font(name="Calibri", size=10, color="FF5E6675")
    ws["F1"].alignment = Alignment(horizontal="right")
    ws.append([])

    columns = ["Rank","Name","Role","Years Exp","Location","Score (Reviewed)","Band","Status",
               "Review Status","Strengths","Gaps","Summary","Flag","Email","Phone","Filename"]
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border

    band_colors = {"strong": GREEN, "possible": ORG, "weak": RED}

    for r in results:
        sc = r.get("overall", 0)
        bname, _, _, _ = band(sc)
        sl = r.get("shortlisted")
        status = "Shortlisted" if sl is True else "Rejected" if sl is False else "Pending"
        ck = candidate_key(r)
        score_fb = score_feedback.get(ck, {})
        reviewed_score = score_fb.get("human_overall", sc) if score_fb else sc
        review_status  = ("Approved" if score_fb.get("approved") else
                          f"Calibrated ({sc}→{reviewed_score})") if score_fb else "Not reviewed"
        row = [r.get("rank",""), r.get("name", r.get("filename","")),
               r.get("role",""), r.get("years",""), r.get("location",""),
               reviewed_score, bname.capitalize(), status, review_status,
               "; ".join(r.get("strengths",[])), "; ".join(r.get("gaps",[])),
               r.get("summary",""), r.get("flag","") or "",
               r.get("email",""), r.get("phone",""), r.get("filename","")]
        ws.append(row)
        data_row = ws.max_row
        for col_idx in range(1, len(columns)+1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            if data_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LGREY)
        band_cell = ws.cell(row=data_row, column=7)
        fc = band_colors.get(bname, DARK)
        band_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        band_cell.fill = PatternFill("solid", fgColor=fc)
        st_cell = ws.cell(row=data_row, column=8)
        if sl is True:
            st_cell.font = Font(name="Calibri", bold=True, color=WHITE)
            st_cell.fill = PatternFill("solid", fgColor=GREEN)
        elif sl is False:
            st_cell.font = Font(name="Calibri", bold=True, color=WHITE)
            st_cell.fill = PatternFill("solid", fgColor=RED)
        rv_cell = ws.cell(row=data_row, column=9)
        if score_fb:
            rv_cell.font = Font(name="Calibri", bold=True,
                                color=GREEN if score_fb.get("approved") else BLUE, size=10)

    col_widths = [6, 22, 22, 10, 16, 14, 10, 12, 18, 40, 30, 50, 35, 24, 16, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    if history:
        wh = wb.create_sheet("Status History")
        wh.append(["Candidate", "Action", "Timestamp"])
        for col_idx in range(1, 4):
            cell = wh.cell(row=1, column=col_idx)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
        all_entries = sorted(
            [e for entries in history.values() for e in entries],
            key=lambda e: e["ts"], reverse=True
        )
        for e in all_entries:
            wh.append([e.get("name",""), e.get("action","").capitalize(), e.get("ts","")])
            row_idx = wh.max_row
            for col_idx in range(1, 4):
                wh.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center")
                wh.cell(row=row_idx, column=col_idx).border = border
        wh.column_dimensions["A"].width = 24
        wh.column_dimensions["B"].width = 16
        wh.column_dimensions["C"].width = 22
        wh.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── API routes ─────────────────────────────────────────────────────────────────

@app.post("/api/parse-jd")
async def parse_jd(file: UploadFile = File(...), qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    try:
        b = await file.read()
        text = parse_bytes(b, file.filename)
        return {"text": text}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")


@app.post("/api/detect-competencies")
async def detect_competencies(jd: str = Form(...), qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    client = anthropic.Anthropic(api_key=get_api_key())
    msg = client.messages.create(
        model="claude-opus-4-8", max_tokens=300,
        messages=[{"role": "user", "content":
            f"Extract exactly 5 key skill competencies from this job description. "
            f"Return ONLY a JSON array of 5 short strings (3-5 words each), nothing else.\n\n{jd}"}]
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("` \n")
    return {"competencies": json.loads(raw)}


@app.post("/api/screen")
async def screen(
    jd: str = Form(...),
    competencies: str = Form(""),
    calibration: str = Form("[]"),
    role_title: str = Form(""),
    files: List[UploadFile] = File(...),
    qs_token: Optional[str] = Cookie(default=None),
):
    _auth.get_current_user(qs_token)
    cvs = {}
    for f in files:
        try:
            cvs[f.filename] = parse_bytes(await f.read(), f.filename)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read {f.filename}: {e}")

    calib = json.loads(calibration)
    past_examples = _auth.get_screening_examples(role_title) if role_title else []

    client = anthropic.Anthropic(api_key=get_api_key())
    msg = client.messages.create(
        model="claude-opus-4-8", max_tokens=16000,
        messages=[{"role": "user", "content": build_prompt(jd, competencies, cvs, calib, past_examples)}]
    )
    truncated = msg.stop_reason == "max_tokens"
    raw = msg.content[0].text if msg.content else ""
    if not raw:
        raise HTTPException(status_code=502, detail="Claude returned an empty response.")
    try:
        results = extract_json(raw)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not parse Claude response: {e}\n\n{raw[:500]}")
    return {"results": results, "truncated": truncated, "past_examples_used": len(past_examples)}


# ── Feedback / learning routes ─────────────────────────────────────────────────

class FeedbackExample(BaseModel):
    candidate_name: str
    ai_score: int
    final_decision: str   # "shortlisted" or "rejected"
    summary: str = ""
    strengths: str = ""
    gaps: str = ""
    recruiter_note: str = ""

class FeedbackBody(BaseModel):
    role_title: str
    examples: List[FeedbackExample]

@app.post("/api/feedback")
async def save_feedback(body: FeedbackBody, qs_token: Optional[str] = Cookie(default=None)):
    user = _auth.get_current_user(qs_token)
    if not body.role_title.strip():
        raise HTTPException(status_code=400, detail="role_title is required to save learning examples")
    _auth.save_screening_examples(
        body.role_title,
        [ex.model_dump() for ex in body.examples],
        user["username"],
    )
    return {"saved": len(body.examples), "role_title": body.role_title}

@app.get("/api/feedback/roles")
async def feedback_roles(qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    return _auth.list_example_roles()

@app.get("/api/feedback/{role_title}")
async def feedback_for_role(role_title: str, qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    return _auth.get_screening_examples(role_title)


class ExportBody(BaseModel):
    results: list
    role_title: str = ""
    history: dict = {}
    score_feedback: dict = {}


@app.post("/api/export/pdf")
async def export_pdf(body: ExportBody, qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    pdf = generate_pdf(body.results, body.role_title)
    filename = f"CV_Screening_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return Response(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/export/excel")
async def export_excel(body: ExportBody, qs_token: Optional[str] = Cookie(default=None)):
    _auth.get_current_user(qs_token)
    xlsx = generate_excel(body.results, body.role_title, body.history, body.score_feedback)
    filename = f"CV_Screening_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Serve frontend ─────────────────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def root():
    return FileResponse("static/index.html")
